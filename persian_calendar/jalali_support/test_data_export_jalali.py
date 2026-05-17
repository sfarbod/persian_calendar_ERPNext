# Copyright (c) 2025, Persian Calendar contributors
# License: MIT

import json
from datetime import date
from io import BytesIO, StringIO

import frappe
from frappe.tests.utils import FrappeTestCase
from frappe.utils import cstr
from openpyxl import load_workbook

from persian_calendar.jalali_support.data_import_export import apply_data_import_export_patches


class TestDataExportJalali(FrappeTestCase):
	@classmethod
	def setUpClass(cls):
		super().setUpClass()
		apply_data_import_export_patches()

	def setUp(self):
		super().setUp()
		frappe.local.response = frappe._dict()
		if getattr(frappe.local, "form_dict", None) is not None:
			frappe.local.form_dict.pop("export_dates_as_jalali", None)

	TEST_USER = "jalali.data.export@test.local"

	def _doc_for_export(self) -> tuple[str, str, str]:
		"""Returns (doctype, name, date_fieldname) with birth_date/posting_date = 2026-04-10."""
		if frappe.db.table_exists("tabJournal Entry"):
			name = frappe.db.get_value("Journal Entry", {}, "name", order_by="modified desc")
			if name:
				frappe.db.set_value(
					"Journal Entry", name, "posting_date", "2026-04-10", update_modified=False
				)
				return "Journal Entry", name, "posting_date"

		email = self.TEST_USER
		if not frappe.db.exists("User", email):
			frappe.get_doc(
				{
					"doctype": "User",
					"email": email,
					"first_name": "Jalali Export Test",
					"enabled": 1,
					"send_welcome_email": 0,
				}
			).insert(ignore_permissions=True)
		frappe.db.set_value("User", email, "birth_date", "2026-04-10", update_modified=False)
		return "User", email, "birth_date"

	def _export_kwargs(self, doctype: str, docname: str, date_field: str, file_type: str = "CSV") -> dict:
		return {
			"doctype": doctype,
			"select_columns": json.dumps({doctype: ["name", date_field]}),
			"filters": json.dumps([[doctype, "name", "=", docname]]),
			"with_data": 1,
			"file_type": file_type,
			"template": True,
			"all_doctypes": False,
			"export_dates_as_jalali": 1,
		}

	@staticmethod
	def _xlsx_as_text(content: bytes) -> str:
		wb = load_workbook(BytesIO(content), data_only=True)
		ws = wb.active
		parts = []
		for row in ws.iter_rows(values_only=True):
			for cell in row:
				if cell is not None and str(cell).strip() != "":
					parts.append(str(cell))
		return "\n".join(parts)

	def test_add_data_row_converts_posting_date(self):
		from frappe.core.doctype.data_export.exporter import DataExporter

		exporter = DataExporter.__new__(DataExporter)
		exporter.export_dates_as_jalali = 1
		exporter.all_doctypes = False
		exporter.columns = ["name", "posting_date"]
		exporter.column_start_end = {
			("Journal Entry", None): frappe._dict(start=0, end=2),
		}
		exporter._jalali_export_debug_rows = 0

		rows = [[""] * 3]
		doc = {"name": "JE-TEST", "posting_date": "2026-04-10"}
		DataExporter.add_data_row(exporter, rows, "Journal Entry", None, doc, 0)

		self.assertEqual(rows[0][2], "1405-01-21")

	def test_real_export_data_csv_contains_jalali(self):
		from frappe.core.doctype.data_export.exporter import export_data

		doctype, docname, date_field = self._doc_for_export()
		export_data(**self._export_kwargs(doctype, docname, date_field, file_type="CSV"))
		result = cstr(frappe.local.response.get("result") or "")
		self.assertIn("1405-01-21", result, msg=result[:2500])

	def test_real_export_data_excel_contains_jalali(self):
		from frappe.core.doctype.data_export.exporter import export_data

		doctype, docname, date_field = self._doc_for_export()
		export_data(**self._export_kwargs(doctype, docname, date_field, file_type="Excel"))
		content = frappe.local.response.get("filecontent")
		self.assertTrue(content, "Excel export should set filecontent on response")
		text = self._xlsx_as_text(content)
		self.assertIn("1405-01-21", text)
		self.assertNotRegex(
			text,
			r"(?<![0-9])10-04-2026(?![0-9])",
			msg="Gregorian dd-mm-yyyy date should not appear in export body",
		)

	def test_real_export_via_frappe_call_excel(self):
		doctype, docname, date_field = self._doc_for_export()
		kwargs = self._export_kwargs(doctype, docname, date_field, file_type="Excel")
		frappe.call("frappe.core.doctype.data_export.exporter.export_data", **kwargs)
		content = frappe.local.response.get("filecontent")
		self.assertTrue(content)
		text = self._xlsx_as_text(content)
		self.assertIn("1405-01-21", text)

	def test_export_data_reads_jalali_from_form_dict(self):
		from frappe.core.doctype.data_export.exporter import export_data

		doctype, docname, date_field = self._doc_for_export()
		kwargs = self._export_kwargs(doctype, docname, date_field, file_type="CSV")
		kwargs.pop("export_dates_as_jalali", None)
		frappe.local.form_dict = frappe._dict({**kwargs, "export_dates_as_jalali": "1"})
		export_data(**kwargs)
		result = cstr(frappe.local.response.get("result") or "")
		self.assertIn("1405-01-21", result)

	def test_export_unchecked_keeps_gregorian_in_data(self):
		from frappe.core.doctype.data_export.exporter import export_data

		doctype, docname, date_field = self._doc_for_export()
		kwargs = self._export_kwargs(doctype, docname, date_field, file_type="CSV")
		kwargs["export_dates_as_jalali"] = 0
		export_data(**kwargs)
		result = cstr(frappe.local.response.get("result") or "")
		self.assertNotIn("1405-01-21", result)
		self.assertTrue("2026" in result or "04" in result or "10" in result)
